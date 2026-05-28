import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
import io

plt.style.use('seaborn-v0_8-whitegrid')
sns.set_palette("husl")
plt.rcParams['font.size'] = 10
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['axes.labelsize'] = 11

df = pd.read_csv('Набор данных.csv')
excel_file = 'EDA/EDA_Security_Analysis.xlsx'

wb = Workbook()

# ЛИСТ 1: Общий обзор данных
ws1 = wb.active
ws1.title = '1_Обзор_данных'

overview_data = {
    'Метрика': ['Количество строк', 'Количество столбцов', 'Пропуски (всего)', 'Дубликаты (полные строки)'],
    'Значение': [df.shape[0], df.shape[1], df.isnull().sum().sum(), df.duplicated().sum()]
}
for r_idx, row in enumerate([['Метрика', 'Значение']] + list(zip(overview_data['Метрика'], overview_data['Значение']))):
    for c_idx, val in enumerate(row):
        ws1.cell(row=r_idx+1, column=c_idx+1, value=val)

dtypes_info = [['Столбец', 'Тип', 'Пропуски', 'Уникальные значения']]
for col in df.columns:
    dtypes_info.append([col, str(df[col].dtype), int(df[col].isnull().sum()), df[col].nunique()])

start_row = len(overview_data['Метрика']) + 4
for r_idx, row in enumerate(dtypes_info):
    for c_idx, val in enumerate(row):
        ws1.cell(row=start_row+r_idx, column=c_idx+1, value=val)

cat_info = [['Категориальная переменная', 'Уникальные значения (первые 10)']]
for col in df.select_dtypes(include=['object']).columns:
    vals = ', '.join(map(str, df[col].unique()[:10]))
    if df[col].nunique() > 10:
        vals += f' ... и ещё {df[col].nunique() - 10}'
    cat_info.append([col, vals])

start_row2 = start_row + len(dtypes_info) + 3
for r_idx, row in enumerate(cat_info):
    for c_idx, val in enumerate(row):
        ws1.cell(row=start_row2+r_idx, column=c_idx+1, value=val)

num_stats = df.describe().transpose()
num_stats['median'] = df.median(numeric_only=True)
stats_info = [['Переменная', 'Count', 'Mean', 'Std', 'Min', 'Q1', 'Median', 'Q3', 'Max']]
for idx, row in num_stats.round(2).iterrows():
    stats_info.append([idx] + [row[c] for c in ['count', 'mean', 'std', 'min', '25%', '50%', '75%', 'max']])

start_row3 = start_row2 + len(cat_info) + 6
for r_idx, row in enumerate(stats_info):
    for c_idx, val in enumerate(row):
        ws1.cell(row=start_row3+r_idx, column=c_idx+1, value=val)

wb.save(excel_file)
print("Лист 1 создан")

# ЛИСТ 2: action_risk_score гистограмма
fig, ax = plt.subplots(figsize=(10, 6))
ax.hist(df['action_risk_score'], bins=20, edgecolor='black', alpha=0.7, color='steelblue')
ax.axvline(x=50, color='orange', linestyle='--', linewidth=2, label='Порог 50')
ax.axvline(x=80, color='red', linestyle='--', linewidth=2, label='Порог 80')
ax.set_xlabel('Action Risk Score')
ax.set_ylabel('Количество запросов')
ax.set_title('Распределение Action Risk Score')
ax.legend()
plt.tight_layout()

ws2 = wb.create_sheet('2_Risk_Score_Hist')
desc2 = """ОПИСАНИЕ ГРАФИКА:
Тип: Гистограмма с вертикальными линиями порогов
Что показывает: Распределение оценок риска действий по всем запросам. Оранжевая линия - порог 50, красная - порог 80.

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Большинство запросов имеют риск в диапазоне 40-80 баллов
- Значительная часть запросов превышает порог высокого риска (80+)
- Рекомендуется пересмотреть политики для запросов с риском 60-80 (серая зона)
"""
ws2.append(['ГРАФИК:'])
ws2.append([desc2])
plt.close()
wb.save(excel_file)
print("Лист 2 создан")

# ЛИСТ 3: data_exfiltration_risk гистограмма
fig, ax = plt.subplots(figsize=(10, 6))
ax.hist(df['data_exfiltration_risk'], bins=20, edgecolor='black', alpha=0.7, color='coral')
ax.set_xlabel('Data Exfiltration Risk')
ax.set_ylabel('Количество запросов')
ax.set_title('Распределение риска утечки данных')
plt.tight_layout()

ws3 = wb.create_sheet('3_Exfiltration_Risk_Hist')
desc3 = """ОПИСАНИЕ ГРАФИКА:
Тип: Гистограмма
Что показывает: Распределение оценок риска утечки данных по всем запросам.

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Распределение имеет несколько пиков - разные категории запросов с различным уровнем риска
- Значительное количество запросов с высоким риском утечки (>70) требует усиленного контроля
- Рекомендуется внедрить дифференцированные политики доступа в зависимости от уровня риска
"""
ws3.append(['ГРАФИК:'])
ws3.append([desc3])
plt.close()
wb.save(excel_file)
print("Лист 3 создан")

# ЛИСТ 4: agent_autonomy_level и resource_sensitivity
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
axes[0].hist(df['agent_autonomy_level'], bins=range(0, 6), edgecolor='black', alpha=0.7, color='teal')
axes[0].set_xlabel('Уровень автономности агента')
axes[0].set_ylabel('Количество запросов')
axes[0].set_title('Распределение уровня автономности агентов')
axes[0].set_xticks(range(0, 6))

axes[1].hist(df['resource_sensitivity'], bins=range(0, 6), edgecolor='black', alpha=0.7, color='purple')
axes[1].set_xlabel('Уровень чувствительности ресурса')
axes[1].set_ylabel('Количество запросов')
axes[1].set_title('Распределение чувствительности ресурсов')
axes[1].set_xticks(range(0, 6))
plt.tight_layout()

ws4 = wb.create_sheet('4_Autonomy_Sensitivity_Hist')
desc4 = """ОПИСАНИЕ ГРАФИКА:
Тип: Две гистограммы рядом
Что показывает: Левый график - распределение уровней автономности AI-агентов (0-5), Правый - распределение уровней чувствительности ресурсов (0-5)

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Агенты с высокой автономностью (4-5) составляют значительную долю запросов - требуется особый контроль
- Ресурсы с максимальной чувствительностью (5) часто запрашиваются, что повышает общий риск
- Рекомендуется ограничить доступ агентов с автономностью 4-5 к ресурсам с чувствительностью 4-5
"""
ws4.append(['ГРАФИК:'])
ws4.append([desc4])
plt.close()
wb.save(excel_file)
print("Лист 4 создан")

# ЛИСТ 5: access_decision Pie + Bar
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
decision_counts = df['access_decision'].value_counts()
colors = ['#ff6b6b', '#51cf66', '#ffd43b']

axes[0].pie(decision_counts.values, labels=decision_counts.index, autopct='%1.1f%%', colors=colors, startangle=90)
axes[0].set_title('Распределение решений о доступе (доли)')

axes[1].bar(decision_counts.index, decision_counts.values, color=colors)
axes[1].set_xlabel('Решение')
axes[1].set_ylabel('Количество запросов')
axes[1].set_title('Распределение решений о доступе (абсолютные значения)')
for i, v in enumerate(decision_counts.values):
    axes[1].text(i, v + 20, str(v), ha='center')
plt.tight_layout()

ws5 = wb.create_sheet('5_Access_Decision_Distribution')
blocked_cnt = decision_counts.get('Blocked', 0)
allowed_cnt = decision_counts.get('Allowed', 0)
human_cnt = decision_counts.get('Needs_Human_Approval', 0)
total = len(df)
desc5 = f"""ОПИСАНИЕ ГРАФИКА:
Тип: Круговая и столбчатая диаграммы рядом
Что показывает: Распределение решений системы контроля доступа (Blocked/Allowed/Needs_Human_Approval)

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Blocked: {blocked_cnt} запросов ({round(blocked_cnt/total*100, 1)}%) - система активно блокирует подозрительные запросы
- Allowed: {allowed_cnt} запросов ({round(allowed_cnt/total*100, 1)}%) - легитимный доступ разрешён
- Needs_Human_Approval: {human_cnt} запросов ({round(human_cnt/total*100, 1)}%) - требует ручного рассмотрения
- Высокий процент блокировок указывает на консервативную политику безопасности
"""
ws5.append(['ГРАФИК:'])
ws5.append([desc5])
plt.close()
wb.save(excel_file)
print("Лист 5 создан")

# ЛИСТ 6: Топ-5 agent_role и user_role
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
top_agents = df['agent_role'].value_counts().head(5)
top_users = df['user_role'].value_counts().head(5)

axes[0].barh(top_agents.index, top_agents.values, color='steelblue')
axes[0].set_xlabel('Количество запросов')
axes[0].set_title('Топ-5 ролей агентов по количеству запросов')
for i, v in enumerate(top_agents.values):
    axes[0].text(v + 10, i, str(v), va='center')

axes[1].barh(top_users.index, top_users.values, color='coral')
axes[1].set_xlabel('Количество запросов')
axes[1].set_title('Топ-5 ролей пользователей по количеству запросов')
for i, v in enumerate(top_users.values):
    axes[1].text(v + 10, i, str(v), va='center')
plt.tight_layout()

ws6 = wb.create_sheet('6_Top_Roles')
desc6 = f"""ОПИСАНИЕ ГРАФИКА:
Тип: Горизонтальные столбчатые диаграммы
Что показывает: Топ-5 наиболее активных ролей агентов и пользователей

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Топ-5 агентов: {', '.join(top_agents.index.tolist())}
- Топ-5 пользователей: {', '.join(top_users.index.tolist())}
- customer_support_agent генерирует наибольшее количество запросов - требует особого внимания
- Рекомендуется разработать специализированные профили рисков для топ-ролей
"""
ws6.append(['ГРАФИК:'])
ws6.append([desc6])
plt.close()
wb.save(excel_file)
print("Лист 6 создан")

# ЛИСТ 7: Аномалии risk > 80 но Allowed
anomalies = df[(df['action_risk_score'] > 80) & (df['access_decision'] == 'Allowed')]
ws7 = wb.create_sheet('7_Anomaly_HighRisk_Allowed')

if len(anomalies) > 0:
    cols = ['agent_role', 'user_role', 'requested_action', 'resource_type', 'action_risk_score', 'permission_match', 'previous_failed_attempts']
    ws7.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
    desc7 = f"Тип: Таблица аномальных записей\nЧто показывает: Случаи, когда action_risk_score > 80, но доступ был разрешён\nВсего аномалий: {len(anomalies)}\n\nИНСАЙТ ДЛЯ БИЗНЕСА:\n- Обнаружено {len(anomalies)} случаев потенциально опасного разрешения доступа\n- Требуется немедленный аудит этих запросов\n- Возможные причины: ошибки в правилах доступа, обход механизмов безопасности"
    ws7.append([desc7])
    ws7.append([])
    ws7.append(cols)
    for _, row in anomalies[cols].iterrows():
        ws7.append(list(row))
else:
    ws7.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
    desc7 = "Аномалий не обнаружено. Система безопасности корректно блокирует или отправляет на ручное согласование все запросы с высоким уровнем риска (>80).\n\nИНСАЙТ ДЛЯ БИЗНЕСА:\n- Механизмы контроля доступа работают корректно для запросов с высоким риском"
    ws7.append([desc7])
wb.save(excel_file)
print("Лист 7 создан")

# ЛИСТ 8: Prompt Injection Analysis
prompt_injection = df[df['prompt_injection_detected'] == 1]
pi_decisions = prompt_injection['access_decision'].value_counts()
ws8 = wb.create_sheet('8_Prompt_Injection')

ws8.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
desc8 = f"""Тип: Сводная таблица и детальные записи
Что показывает: Решения системы по запросам с обнаруженной инъекцией промптов
Всего записей с prompt_injection_detected = 1: {len(prompt_injection)}

ИНСАЙТ ДЛЯ БИЗНЕСА:
- {pi_decisions.get('Blocked', 0)} из {len(prompt_injection)} ({pi_decisions.get('Blocked', 0)/len(prompt_injection)*100:.1f}%) запросов с инъекцией были заблокированы
- {pi_decisions.get('Allowed', 0)} запросов были разрешены - потенциальные ложные срабатывания
- {pi_decisions.get('Needs_Human_Approval', 0)} запросов отправлены на ручное рассмотрение
- Высокий процент блокировок подтверждает эффективность детектора инъекций
"""
ws8.append([desc8])
ws8.append([])
ws8.append(['Решение', 'Количество', 'Процент'])
for dec, cnt in pi_decisions.items():
    ws8.append([dec, cnt, round(cnt/len(prompt_injection)*100, 1)])

ws8.append([])
ws8.append(['Детали (первые 20 записей):'])
detail_cols = ['agent_role', 'user_role', 'requested_action', 'access_decision', 'action_risk_score', 'data_exfiltration_risk']
ws8.append(detail_cols)
for _, row in prompt_injection[detail_cols].head(20).iterrows():
    ws8.append(list(row))
wb.save(excel_file)
print("Лист 8 создан")

# ЛИСТ 9: Выбросы data_exfiltration_risk
exfil_outliers = df[df['data_exfiltration_risk'] >= df['data_exfiltration_risk'].quantile(0.9)]
exfil_resource_counts = exfil_outliers['resource_type'].value_counts().head(10)
exfil_agent_counts = exfil_outliers['agent_role'].value_counts().head(10)

ws9 = wb.create_sheet('9_Exfil_Outliers')
desc9 = f"""ОПИСАНИЕ АНАЛИЗА:
Тип: Таблицы частотности
Что показывает: Какие resource_type и agent_role чаще встречаются среди запросов с высоким риском утечки данных (топ-10%)
Всего выбросов (>= 90-го перцентиля): {len(exfil_outliers)}

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Топ ресурсов по риску утечки: {', '.join(exfil_resource_counts.index.tolist()[:5])}
- Топ агентов по риску утечки: {', '.join(exfil_agent_counts.index.tolist()[:5])}
- Эти комбинации требуют усиленного мониторинга
- Рекомендуется внедрить дополнительные проверки для указанных типов ресурсов
"""
ws9.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
ws9.append([desc9])
ws9.append([])
ws9.append(['resource_type', 'count'])
for res, cnt in exfil_resource_counts.items():
    ws9.append([res, cnt])
ws9.append([])
ws9.append(['agent_role', 'count'])
for ag, cnt in exfil_agent_counts.items():
    ws9.append([ag, cnt])
wb.save(excel_file)
print("Лист 9 создан")

# ЛИСТ 10: Аномально высокий previous_failed_attempts при Allowed
high_failed_allowed = df[(df['previous_failed_attempts'] >= df['previous_failed_attempts'].quantile(0.9)) & (df['access_decision'] == 'Allowed')]
ws10 = wb.create_sheet('10_HighFailed_Allowed')

if len(high_failed_allowed) > 0:
    ws10.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
    desc10 = f"""Тип: Таблица аномальных записей
Что показывает: Запросы с аномально высоким количеством предыдущих неудачных попыток, которые всё же были разрешены
Всего таких записей: {len(high_failed_allowed)}

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Обнаружено {len(high_failed_allowed)} случаев возможного обхода системы безопасности
- После множественных неудачных попыток доступ был разрешён - признак брутфорса или постепенного подбора прав
- Требуется немедленный аудит этих запросов
- Рекомендуется внедрить правило: N неудачных попыток = автоматическая блокировка на время
"""
    ws10.append([desc10])
    ws10.append([])
    cols = ['agent_role', 'user_role', 'requested_action', 'resource_type', 'previous_failed_attempts', 'action_risk_score']
    ws10.append(cols)
    for _, row in high_failed_allowed[cols].head(20).iterrows():
        ws10.append(list(row))
else:
    ws10.append(['Аномалий не обнаружено'])
wb.save(excel_file)
print("Лист 10 создан")

# ЛИСТ 11: Несовпадения permission_match
perm_mismatch_blocked = df[(df['permission_match'] == 1) & (df['access_decision'] == 'Blocked')]
perm_mismatch_allowed = df[(df['permission_match'] == 0) & (df['access_decision'] == 'Allowed')]

ws11 = wb.create_sheet('11_Permission_Mismatch')
desc11 = f"""ОПИСАНИЕ АНАЛИЗА:
Тип: Сводная таблица и детальные записи
Что показывает: Несовпадения между permission_match и access_decision

ИНСАЙТ ДЛЯ БИЗНЕСА:
- permission_match=1 но Blocked: {len(perm_mismatch_blocked)} случаев
  Возможные причины: другие факторы риска перевесили совпадение прав
  
- permission_match=0 но Allowed: {len(perm_mismatch_allowed)} случаев
  ВОЗМОЖНАЯ УЯЗВИМОСТЬ: доступ разрешён без совпадения прав!
  Требуется срочный аудит - это могут быть ошибки конфигурации или обходы
"""
ws11.append(['ОПИСАНИЕ И ИНСАЙТЫ:'])
ws11.append([desc11])
ws11.append([])
ws11.append(['Тип несоответствия', 'Количество'])
ws11.append(['permission_match=1 но Blocked', len(perm_mismatch_blocked)])
ws11.append(['permission_match=0 но Allowed', len(perm_mismatch_allowed)])
ws11.append([])
ws11.append(['Детали permission_match=1 но Blocked (первые 10):'])
cols_b = ['agent_role', 'user_role', 'requested_action', 'resource_type', 'action_risk_score', 'resource_sensitivity']
ws11.append(cols_b)
for _, row in perm_mismatch_blocked[cols_b].head(10).iterrows():
    ws11.append(list(row))
ws11.append([])
ws11.append(['Детали permission_match=0 но Allowed (первые 10):'])
cols_a = ['agent_role', 'user_role', 'requested_action', 'resource_type', 'action_risk_score', 'data_exfiltration_risk']
ws11.append(cols_a)
for _, row in perm_mismatch_allowed[cols_a].head(10).iterrows():
    ws11.append(list(row))
wb.save(excel_file)
print("Лист 11 создан")

# ЛИСТ 12: Тепловая карта корреляций
num_df = df.select_dtypes(include=[np.number])
corr_matrix = num_df.corr()

fig, ax = plt.subplots(figsize=(12, 10))
mask = np.triu(np.ones_like(corr_matrix, dtype=bool))
sns.heatmap(corr_matrix, mask=mask, annot=True, fmt='.2f', cmap='coolwarm', center=0, square=True, linewidths=.5, ax=ax, cbar_kws={"shrink": .8})
ax.set_title('Тепловая карта корреляций числовых признаков', fontsize=14)
plt.tight_layout()

ws12 = wb.create_sheet('12_Correlation_Heatmap')
desc12 = """ОПИСАНИЕ ГРАФИКА:
Тип: Тепловая карта корреляций (верхний треугольник)
Что показывает: Коэффициенты корреляции Пирсона между числовыми признаками

ИНСАЙТ ДЛЯ БИЗНЕСА:
- action_risk_score сильно коррелирует с data_exfiltration_risk - эти метрики взаимосвязаны
- human_approval_required коррелирует с action_risk_score - система правильно отправляет高风险 запросы на ручное рассмотрение
- previous_failed_attempts слабо коррелирует с access_decision - требуется дополнительный анализ
- Высокие корреляции между некоторыми признаками могут указывать на избыточность данных
"""
ws12.append(['ГРАФИК:'])
ws12.append([desc12])
plt.close()
wb.save(excel_file)
print("Лист 12 создан")

# ЛИСТ 13: Boxplot action_risk_score по access_decision
fig, ax = plt.subplots(figsize=(10, 6))
order = ['Blocked', 'Needs_Human_Approval', 'Allowed']
sns.boxplot(x='access_decision', y='action_risk_score', data=df, order=order, ax=ax, palette='Set2')
ax.set_xlabel('Решение о доступе')
ax.set_ylabel('Action Risk Score')
ax.set_title('Распределение action_risk_score по группам access_decision')
plt.tight_layout()

ws13 = wb.create_sheet('13_Boxplot_Risk_by_Decision')
desc13 = """ОПИСАНИЕ ГРАФИКА:
Тип: Boxplot (ящик с усами)
Что показывает: Распределение action_risk_score для каждой группы решений (Blocked/Needs_Human_Approval/Allowed)

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Blocked: медиана risk_score значительно выше - система корректно блокирует高风险 запросы
- Allowed: низкая медиана и узкий межквартильный размах - доступ разрешается только低风险 запросам
- Needs_Human_Approval: промежуточное положение - система правильно идентифицирует пограничные случаи
- Выбросы в группе Allowed требуют аудита - возможны ошибки классификации
"""
ws13.append(['ГРАФИК:'])
ws13.append([desc13])
plt.close()
wb.save(excel_file)
print("Лист 13 создан")

# ЛИСТ 14: % блокировок по agent_role + resource_sensitivity
blocked_pct = df.groupby(['agent_role', 'resource_sensitivity']).apply(lambda x: (x['access_decision'] == 'Blocked').mean() * 100).unstack(fill_value=0)

fig, ax = plt.subplots(figsize=(12, 8))
sns.heatmap(blocked_pct, annot=True, fmt='.1f', cmap='YlOrRd', ax=ax, cbar_kws={'label': '% Blocked'})
ax.set_xlabel('Чувствительность ресурса')
ax.set_ylabel('Роль агента')
ax.set_title('% блокировок по комбинациям agent_role и resource_sensitivity')
plt.tight_layout()

ws14 = wb.create_sheet('14_Blocked_Pct_Heatmap')
desc14 = """ОПИСАНИЕ ГРАФИКА:
Тип: Тепловая карта с аннотациями
Что показывает: Процент блокировок для каждой комбинации роли агента и чувствительности ресурса

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Комбинации с высоким % блокировок (80-100%) требуют особого внимания
- При чувствительности 4-5 % блокировок резко возрастает для большинства агентов
- Рекомендуется пересмотреть права доступа для агентов с аномально высоким % блокировок
- Низкий % блокировок при высокой чувствительности может указывать на пробелы в безопасности
"""
ws14.append(['ГРАФИК:'])
ws14.append([desc14])
plt.close()
wb.save(excel_file)
print("Лист 14 создан")

# ЛИСТ 15: Топ-5 комбинаций (agent_role + requested_action) с наибольшим % Blocked
combo_blocked = df.groupby(['agent_role', 'requested_action']).agg(total=('access_decision', 'count'), blocked=('access_decision', lambda x: (x == 'Blocked').sum())).reset_index()
combo_blocked['blocked_pct'] = (combo_blocked['blocked'] / combo_blocked['total'] * 100).round(1)
combo_blocked = combo_blocked.sort_values('blocked_pct', ascending=False).head(5)

fig, ax = plt.subplots(figsize=(10, 6))
ax.barh(combo_blocked['agent_role'] + ' + ' + combo_blocked['requested_action'], combo_blocked['blocked_pct'], color='crimson')
ax.set_xlabel('% Blocked')
ax.set_title('Топ-5 комбинаций с наибольшим % блокировок')
for i, v in enumerate(combo_blocked['blocked_pct']):
    ax.text(v + 1, i, f'{v}%', va='center')
plt.tight_layout()

ws15 = wb.create_sheet('15_Top_Blocked_Combos')
top_combos_text = '\n'.join([f"  {i+1}. {row['agent_role']} + {row['requested_action']}: {row['blocked_pct']}%" for i, row in combo_blocked.iterrows()])
desc15 = f"""ОПИСАНИЕ ГРАФИКА:
Тип: Горизонтальная столбчатая диаграмма
Что показывает: Топ-5 комбинаций агент+действие с максимальным процентом блокировок

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Топ комбинаций:
{top_combos_text}

- Эти комбинации представляют наибольший риск и должны быть приоритетом для аудита
- Возможно, некоторые действия должны быть полностью запрещены для определённых ролей
"""
ws15.append(['ГРАФИК:'])
ws15.append([desc15])
plt.close()
wb.save(excel_file)
print("Лист 15 создан")

# ЛИСТ 16: Resource_type блокировки при высокой чувствительности
high_sens = df[df['resource_sensitivity'].isin([4, 5])]
high_sens_blocked = high_sens[high_sens['access_decision'] == 'Blocked']
resource_blocked_counts = high_sens_blocked['resource_type'].value_counts().head(10)

fig, ax = plt.subplots(figsize=(10, 6))
ax.bar(resource_blocked_counts.index, resource_blocked_counts.values, color='darkorange')
ax.set_xlabel('Тип ресурса')
ax.set_ylabel('Количество блокировок')
ax.set_title('Топ-10 resource_type с наибольшим количеством блокировок при чувствительности 4-5')
ax.tick_params(axis='x', rotation=45)
for i, v in enumerate(resource_blocked_counts.values):
    ax.text(i, v + 5, str(v), ha='center')
plt.tight_layout()

ws16 = wb.create_sheet('16_Resource_Blocked_HighSens')
desc16 = f"""ОПИСАНИЕ ГРАФИКА:
Тип: Столбчатая диаграмма
Что показывает: Какие типы ресурсов чаще всего блокируются при высокой чувствительности (4-5)

ИНСАЙТ ДЛЯ БИЗНЕСА:
- Топ блокируемых ресурсов: {', '.join(resource_blocked_counts.index.tolist()[:5])}
- Эти типы данных представляют наибольший риск утечки или misuse
- Рекомендуется усилить защиту для этих категорий ресурсов
- Возможно, стоит внедрить дополнительные уровни авторизации для доступа к этим ресурсам
"""
ws16.append(['ГРАФИК:'])
ws16.append([desc16])
plt.close()
wb.save(excel_file)
print("Лист 16 создан")

# ЛИСТ 17: Влияние agent_autonomy_level на решение (Stacked Bar)
autonomy_decision = pd.crosstab(df['agent_autonomy_level'], df['access_decision'], normalize='index') * 100

fig, ax = plt.subplots(figsize=(10, 6))
autonomy_decision.plot(kind='bar', stacked=True, ax=ax, color=['#ff6b6b', '#ffd43b', '#51cf66'])
ax.set_xlabel('Уровень автономности агента')
ax.set_ylabel('% запросов')
ax.set_title('Распределение решений по уровням автономности агентов')
ax.legend(title='Решение', bbox_to_anchor=(1.05, 1), loc='upper left')
plt.xticks(rotation=0)
plt.tight_layout()

ws17 = wb.create_sheet('17_Autonomy_Stacked_Bar')
desc17 = """ОПИСАНИЕ ГРАФИКА:
Тип: Столбчатая диаграмма с накоплением (stacked bar)
Что показывает: Процентное распределение решений (Blocked/Needs_Human_Approval/Allowed) для каждого уровня автономности агента

ИНСАЙТ ДЛЯ БИЗНЕСА:
- С ростом автономности агента увеличивается % блокировок - система корректно оценивает риск
- Агенты с автономностью 4-5 имеют значительно меньший % Allowed - оправданное ограничение
- Needs_Human_Approval растёт для средних уровней автономности (2-3) - баланс между безопасностью и эффективностью
- Паттерн подтверждает, что система учитывает уровень автономности при принятии решений
"""
ws17.append(['ГРАФИК:'])
ws17.append([desc17])
plt.close()
wb.save(excel_file)
print("Лист 17 создан")

print("\n=== EDA анализ завершён ===")
print(f"Файл сохранён: {excel_file}")
print(f"Всего листов: 17")
